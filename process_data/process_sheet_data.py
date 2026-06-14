import argparse
import json
import os
import random
import re
import shutil
import string
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_FILE_PATH = Path(__file__).resolve().parents[1] / "data" / "san-pham-shopee-resell.xlsx"
DEFAULT_OUTPUT_FILE_PATH = Path(__file__).resolve().parents[1] / "data" / "copy_data.xlsx"
_REPO_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_KEY_CANDIDATES = (
    _REPO_ROOT / "bigseller-tools" / "openai.key",
    Path(__file__).resolve().parent / "openai_api_key.txt",
)
DEFAULT_API_KEY_FILE_PATH = next(
    (path for path in _DEFAULT_KEY_CANDIDATES if path.exists()),
    _DEFAULT_KEY_CANDIDATES[-1],
)
SAVE_EVERY_UNIQUE_NAMES = 50
COPY_PREFIX = "copy_"
DEFAULT_MODEL = "gpt-4o-mini"
PRODUCT_NAME_HEADER = "Tên sp"
SKU_HEADER = "SKU"
PRODUCT_LINK_HEADER = "link sp"
ATTRIBUTE_START_PATTERNS = [
    ("cao",),
    ("êm", "chân"),
    ("da", "mềm"),
    ("đính", "nơ"),
    ("dây", "cài"),
    ("quai", "cài"),
    ("may", "viền"),
    ("hot", "trend"),
    ("dễ", "phối"),
    ("tôn", "dáng"),
    ("phong", "cách"),
]


def normalize_sheet_name(sheet_name):
    name = str(sheet_name).strip().lower()
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"[^a-z0-9_]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name or "sheet"


def build_copy_sheet_name(source_sheet_name):
    # Excel sheet names are limited to 31 characters.
    return f"{COPY_PREFIX}{normalize_sheet_name(source_sheet_name)}"[:31]


def get_unique_sheet_name(workbook, preferred_name):
    if preferred_name not in workbook.sheetnames:
        return preferred_name

    base_name = preferred_name[:28]
    index = 1
    while True:
        candidate = f"{base_name}_{index}"[:31]
        if candidate not in workbook.sheetnames:
            return candidate
        index += 1


def get_source_sheet_names(workbook, template_sheet_name):
    return [
        sheet_name
        for sheet_name in workbook.sheetnames
        if sheet_name != template_sheet_name
        and not sheet_name.lower().startswith(COPY_PREFIX)
    ]


def get_template_sheet_name(workbook):
    non_copy_sheet_names = [
        sheet_name for sheet_name in workbook.sheetnames if not sheet_name.lower().startswith(COPY_PREFIX)
    ]
    if not non_copy_sheet_names:
        raise ValueError("Workbook does not contain a non-copy template sheet.")
    return non_copy_sheet_names[-1]


def find_header_column(sheet, header_name):
    normalized_header_name = str(header_name).strip().lower()

    for cell in sheet[1]:
        if str(cell.value).strip().lower() == normalized_header_name:
            return cell.column

    return None


def get_existing_skus(workbook):
    existing_skus = set()

    for sheet in workbook.worksheets:
        sku_column = find_header_column(sheet, SKU_HEADER)
        if not sku_column:
            continue

        for row_index in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=sku_column).value
            if value is not None and str(value).strip():
                existing_skus.add(str(value).strip())

    return existing_skus


def generate_unique_sku(existing_skus):
    while True:
        sku = "B" + "".join(random.choices(string.digits, k=5))
        if sku not in existing_skus:
            existing_skus.add(sku)
            return sku


def update_missing_skus(sheet, existing_skus):
    sku_column = find_header_column(sheet, SKU_HEADER)
    product_name_column = find_header_column(sheet, PRODUCT_NAME_HEADER)
    updated_count = 0

    if not sku_column or not product_name_column:
        return updated_count

    for row_index in range(2, sheet.max_row + 1):
        has_product_name = (
            sheet.cell(row=row_index, column=product_name_column).value is not None
            and str(sheet.cell(row=row_index, column=product_name_column).value).strip()
        )

        if not has_product_name:
            continue

        sku_cell = sheet.cell(row=row_index, column=sku_column)
        if sku_cell.value is None or not str(sku_cell.value).strip():
            sku_cell.value = generate_unique_sku(existing_skus)
            updated_count += 1

    return updated_count


def normalize_dash(text):
    return re.sub(r"\s*[–—-]\s*", " - ", str(text).strip())


def split_name_code(product_name):
    match = re.search(r"\s+-\s+([A-Z]\d+)\s*$", normalize_dash(product_name))
    if not match:
        return normalize_dash(product_name), None

    body = normalize_dash(product_name)[: match.start()].strip()
    return body, match.group(1)


def starts_with_pattern(words, index, pattern):
    lowered_words = [word.lower().strip(" ,.;:") for word in words[index : index + len(pattern)]]
    return tuple(lowered_words) == pattern


def infer_locked_context(product_name):
    body, code = split_name_code(product_name)
    parts = body.split(" - ", 1)

    if len(parts) == 1:
        locked_context = body
    else:
        first_part, second_part = parts
        second_words = second_part.split()
        stop_index = len(second_words)

        for index in range(2, len(second_words)):
            if any(starts_with_pattern(second_words, index, pattern) for pattern in ATTRIBUTE_START_PATTERNS):
                stop_index = index
                break

        locked_second_part = " ".join(second_words[:stop_index]).strip()
        locked_context = f"{first_part.strip()} - {locked_second_part}".strip()

    return locked_context, code


def infer_product_name_structure(product_name):
    locked_context, code = infer_locked_context(product_name)
    parts = locked_context.split(" - ", 1)

    if len(parts) == 2:
        keyword_1, keyword_2 = parts
    else:
        keyword_1 = locked_context
        keyword_2 = ""

    body, _ = split_name_code(product_name)
    description = body
    if keyword_2:
        prefix = f"{keyword_1} - {keyword_2}"
        if body.startswith(prefix):
            description = body[len(prefix):].strip()
    elif body.startswith(keyword_1):
        description = body[len(keyword_1):].strip()

    return {
        "keyword_1": keyword_1.strip(),
        "keyword_2": keyword_2.strip(),
        "description": description.strip(" -"),
        "product_code": code or "",
    }


def normalize_product_name_structure(original_name, structure):
    fallback = infer_product_name_structure(original_name)
    if not isinstance(structure, dict):
        return fallback

    keyword_1 = str(structure.get("keyword_1", "")).strip() or fallback["keyword_1"]
    keyword_2 = str(structure.get("keyword_2", "")).strip() or fallback["keyword_2"]
    description = str(structure.get("description", "")).strip() or fallback["description"]
    product_code = str(structure.get("product_code", "")).strip() or fallback["product_code"]

    return {
        "keyword_1": keyword_1,
        "keyword_2": keyword_2,
        "description": description,
        "product_code": product_code,
    }


def compose_product_name(structure, description=None):
    keyword_1 = str(structure.get("keyword_1", "")).strip()
    keyword_2 = str(structure.get("keyword_2", "")).strip()
    product_code = str(structure.get("product_code", "")).strip()
    final_description = str(description if description is not None else structure.get("description", "")).strip()

    parts = [keyword_1]
    if keyword_2:
        parts.append(keyword_2)
    body = " - ".join(part for part in parts if part).strip()
    if final_description:
        body = f"{body} {final_description}".strip()
    if product_code:
        return f"{body} - {product_code}"
    return body


def extract_response_text(response_body):
    if response_body.get("output_text"):
        return response_body["output_text"]

    output_parts = []
    for item in response_body.get("output", []):
        for content in item.get("content", []):
            if content.get("type") == "output_text" and content.get("text"):
                output_parts.append(content["text"])

    return "\n".join(output_parts)


def describe_http_error(error):
    try:
        error_body = error.read().decode("utf-8")
    except Exception:
        error_body = str(error)

    return f"{error.code} {error.reason}: {error_body}"


def build_indexed_product_names(product_names):
    return [
        {
            "index": index,
            "name": product_name,
        }
        for index, product_name in enumerate(product_names)
    ]


def request_product_name_structures(product_names, model, api_key, max_retries=3):
    payload = {
        "model": model,
        "instructions": (
            "Toi se gui danh sach ten san pham bang tieng Viet. "
            "Voi moi ten, hay tach thanh 4 truong: keyword_1, keyword_2, description, product_code. "
            "keyword_1 la ve truoc dau '-' hoac '–'. "
            "keyword_2 la cum loai san pham chinh ngay sau dau '-' hoac '–', phai ngan gon. "
            "keyword_2 thuong chi la 2 den 4 tu, vi du: 'Giay Bup Be', 'Dep Suc Nu', 'Giay Mary Jane'. "
            "Khong duoc dua chat lieu, hoa tiet, dac diem, cong dung, ngu canh su dung vao keyword_2. "
            "Tat ca phan con lai sau keyword_2 phai dua vao description. "
            "description la phan mo ta con lai sau keyword_2, gom chat lieu, hoa tiet, kieu dang, dac diem, ngu canh su dung neu co. "
            "product_code la ma cuoi ten nhu A614, A696, A745 neu co. "
            "Vi du: 'Giay Bet Nu - Giay Bup Be Luoi Ren Dinh Da Thoang Khi Nhe Nhang Nu Tinh Em Ai De Phoi Do Hang Ngay - B91763' "
            "thi keyword_1='Giay Bet Nu', keyword_2='Giay Bup Be', "
            "description='Luoi Ren Dinh Da Thoang Khi Nhe Nhang Nu Tinh Em Ai De Phoi Do Hang Ngay', product_code='B91763'. "
            "Khong doi nghia, khong viet lai, chi phan tich cau truc. "
            "Moi item output phai giu dung index cua item input tuong ung. "
            "Chi tra ve JSON dung schema."
        ),
        "input": json.dumps(
            {
                "products": build_indexed_product_names(product_names),
            },
            ensure_ascii=False,
        ),
        "text": {
            "format": {
                "type": "json_schema",
                "name": "parsed_product_name_structures",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "index": {"type": "integer"},
                                    "keyword_1": {"type": "string"},
                                    "keyword_2": {"type": "string"},
                                    "description": {"type": "string"},
                                    "product_code": {"type": "string"},
                                },
                                "required": ["index", "keyword_1", "keyword_2", "description", "product_code"],
                            },
                        }
                    },
                    "required": ["items"],
                },
            }
        },
    }

    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    for attempt in range(1, max_retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                response_body = json.loads(response.read().decode("utf-8"))
            response_text = extract_response_text(response_body)
            parsed = json.loads(response_text)
            items = parsed["items"]
            items_by_index = {item["index"]: item for item in items}
            missing_indexes = [index for index in range(len(product_names)) if index not in items_by_index]
            if missing_indexes or len(items_by_index) != len(product_names):
                raise ValueError(
                    "OpenAI returned mismatched parsed indexes. "
                    f"Missing indexes: {missing_indexes}."
                )

            structures = []
            for index, product_name in enumerate(product_names):
                structures.append(normalize_product_name_structure(product_name, items_by_index[index]))
            return structures
        except urllib.error.HTTPError as error:
            message = describe_http_error(error)
            if attempt == max_retries:
                raise RuntimeError(f"OpenAI parse request failed: {message}") from error
            time.sleep(2 * attempt)
        except (urllib.error.URLError, TimeoutError, ValueError, json.JSONDecodeError) as error:
            if attempt == max_retries:
                raise RuntimeError(f"OpenAI parse request failed: {error}") from error
            time.sleep(2 * attempt)

    return [infer_product_name_structure(product_name) for product_name in product_names]


def request_product_name_structures_with_split(product_names, model, api_key):
    try:
        return request_product_name_structures(product_names, model, api_key)
    except RuntimeError:
        if len(product_names) <= 1:
            return [infer_product_name_structure(product_names[0])]

        middle_index = len(product_names) // 2
        left_structures = request_product_name_structures_with_split(
            product_names[:middle_index],
            model,
            api_key,
        )
        right_structures = request_product_name_structures_with_split(
            product_names[middle_index:],
            model,
            api_key,
        )
        return left_structures + right_structures


def request_rewritten_name_versions(parsed_products, version_count, model, api_key, max_retries=3):
    payload = {
        "model": model,
        "instructions": (
            "Toi se gui danh sach san pham da duoc tach cau truc thanh keyword_1, keyword_2, description, product_code. "
            "Hay viet lai duy nhat phan description. "
            "Khong doi keyword_1, keyword_2, product_code. "
            f"Voi moi san pham, tao dung {version_count} phien ban description khac nhau. "
            "Description moi phai ngan gon, dang title san pham, khong phai cau marketing dai. "
            "Description moi can ngan gon, mo ta dung dac diem san pham, uu tien trong khoang 8 den 16 tu. "
            "Do dai nen xap xi description goc, khoang 70 den 100 phan tram, khong dai hon ro ret. "
            "Chi viet mot cum lien tuc, khong tach thanh 2 cau, khong dung dau cham cuoi cau. "
            "Khong lap lai keyword_1 hoac keyword_2 trong description moi. "
            "Khong bat dau description bang cac tu nhu 'giay', 'doi giay', 'dep', 'sandal', 'boots'. "
            "Bam sat y nghia cua description goc, chi thay cach dien dat, khong tu y them loi ich moi hoac slogan moi. "
            "Chi duoc paraphrase description goc bang cach doi cum tu dong nghia va sap xep lai y. "
            "Cam sang tac them y moi, cam them dac diem moi, cam them loi ich moi, cam them ngu canh moi khong co trong description goc. "
            "Neu description goc khong co thong tin do, tuyet doi khong duoc tu them. "
            "Khong dung cac cum tu chung chung nhu 'mang den', 'lua chon hoan hao', 'giup ban', 'noi bat trong moi khong gian', "
            "'phong cach chuyen nghiep', 'de dang ket hop trang phuc', 'phu hop moi dip', 'ly tuong cho', 'hoan hao cho'. "
            "Khong dung mo ta rong, slogan, loi khen chung chung, hoac cau mang tinh quang cao. "
            "Khong dung cac tu cam tinh hoac suy dien them nhu 'quyen ru', 'thu hut', 'noi bat', 'an tuong', 'hoan hao', 'sang chanh', 'dang cap', 'ly tuong', 'dep mat' neu cac y do khong co san trong description goc. "
            "Uu tien tu khoa dac ta truc tiep dac diem san pham, giu chat listing title thuong mai dien tu. "
            "Khong thay doi y nghia san pham, khong bia chat lieu/tinh nang moi. "
            "Cac version cua cung mot san pham phai khac cach dien dat ro rang. "
            "Moi item output phai giu dung index cua item input tuong ung. "
            "Chi tra ve JSON dung schema."
        ),
        "input": json.dumps(
            {
                "version_count": version_count,
                "products": parsed_products,
            },
            ensure_ascii=False,
        ),
        "text": {
            "format": {
                "type": "json_schema",
                "name": "rewritten_product_name_versions",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "index": {"type": "integer"},
                                    "versions": {
                                        "type": "array",
                                        "items": {"type": "string"},
                                    }
                                },
                                "required": ["index", "versions"],
                            },
                        }
                    },
                    "required": ["items"],
                },
            }
        },
    }

    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    for attempt in range(1, max_retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                response_body = json.loads(response.read().decode("utf-8"))
            response_text = extract_response_text(response_body)
            parsed = json.loads(response_text)
            items = parsed["items"]
            items_by_index = {item["index"]: item for item in items}
            missing_indexes = [
                index for index in range(len(parsed_products)) if index not in items_by_index
            ]
            if missing_indexes or len(items_by_index) != len(parsed_products):
                raise ValueError(
                    "OpenAI returned mismatched product indexes. "
                    f"Missing indexes: {missing_indexes}."
                )

            all_versions = []
            for index, parsed_product in enumerate(parsed_products):
                item = items_by_index[index]
                versions = item["versions"]
                if len(versions) != version_count:
                    raise ValueError(
                        f"OpenAI returned {len(versions)} versions for item {index}, expected {version_count}."
                    )
                all_versions.append(
                    [compose_product_name(parsed_product, rewritten_name) for rewritten_name in versions]
                )
            return all_versions
        except urllib.error.HTTPError as error:
            message = describe_http_error(error)
            if attempt == max_retries:
                raise RuntimeError(f"OpenAI request failed: {message}") from error
            time.sleep(2 * attempt)
        except (urllib.error.URLError, TimeoutError, ValueError, json.JSONDecodeError) as error:
            if attempt == max_retries:
                raise RuntimeError(f"OpenAI request failed: {error}") from error
            time.sleep(2 * attempt)

    return [[compose_product_name(parsed_product)] * version_count for parsed_product in parsed_products]


def request_rewritten_name_versions_with_split(parsed_products, version_count, model, api_key):
    try:
        return request_rewritten_name_versions(parsed_products, version_count, model, api_key)
    except RuntimeError:
        if len(parsed_products) <= 1:
            raise

        middle_index = len(parsed_products) // 2
        left_versions = request_rewritten_name_versions_with_split(
            parsed_products[:middle_index],
            version_count,
            model,
            api_key,
        )
        right_versions = request_rewritten_name_versions_with_split(
            parsed_products[middle_index:],
            version_count,
            model,
            api_key,
        )
        return left_versions + right_versions


def get_product_name_rows(sheet):
    product_name_column = find_header_column(sheet, PRODUCT_NAME_HEADER)
    if not product_name_column:
        return []

    rows_to_update = []
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=product_name_column)
        if cell.value is not None and str(cell.value).strip():
            rows_to_update.append((row_index, product_name_column, str(cell.value).strip()))

    return rows_to_update


def remove_copy_sheets(workbook):
    deleted_sheets = []

    for sheet_name in list(workbook.sheetnames):
        if sheet_name.lower().startswith(COPY_PREFIX):
            workbook.remove(workbook[sheet_name])
            deleted_sheets.append(sheet_name)

    return deleted_sheets


def _parse_api_key_line(line):
    line = str(line or "").strip()
    if not line or line.startswith("#"):
        return None
    if "=" in line and not line.startswith("sk-"):
        _, _, line = line.partition("=")
        line = line.strip().strip('"').strip("'")
    return line or None


def get_openai_api_key(api_key_file_path=None):
    env_api_key = os.environ.get("OPENAI_API_KEY")
    if env_api_key:
        for line in env_api_key.splitlines():
            parsed = _parse_api_key_line(line)
            if parsed:
                return parsed

    if api_key_file_path:
        path = Path(api_key_file_path)
        if path.exists():
            for line in path.read_text(encoding="utf-8").splitlines():
                parsed = _parse_api_key_line(line)
                if parsed:
                    return parsed

    return None


def get_progress_file_path(output_file_path):
    return output_file_path.with_suffix(".progress.json")


def load_progress(progress_file_path):
    if not progress_file_path.exists():
        return {"processed_product_names": [], "parsed_structures": {}}

    return json.loads(progress_file_path.read_text(encoding="utf-8"))


def save_progress(progress_file_path, processed_product_names, parsed_structures):
    progress_file_path.write_text(
        json.dumps(
            {
                "processed_product_names": sorted(processed_product_names),
                "parsed_structures": parsed_structures,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def get_copy_sheet_names_from_source_sheet_names(source_sheet_names):
    copy_sheet_names = []
    seen_names = set()

    for source_sheet_name in source_sheet_names:
        preferred_name = build_copy_sheet_name(source_sheet_name)
        target_name = preferred_name
        if target_name in seen_names:
            base_name = preferred_name[:28]
            index = 1
            while True:
                candidate = f"{base_name}_{index}"[:31]
                if candidate not in seen_names:
                    target_name = candidate
                    break
                index += 1
        seen_names.add(target_name)
        copy_sheet_names.append(target_name)

    return copy_sheet_names


def create_initial_copy_workbook(source_file_path, output_file_path):
    if output_file_path.suffix.lower() != ".xlsx":
        raise ValueError("Output file must use .xlsx because openpyxl does not write legacy .xls files.")

    if source_file_path.resolve() == output_file_path.resolve():
        raise ValueError("Output file must be different from the source file.")

    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_file_path, output_file_path)

    workbook = load_workbook(output_file_path)
    deleted_sheets = remove_copy_sheets(workbook)
    template_sheet = workbook.worksheets[-1]
    template_sheet_name = template_sheet.title
    source_sheet_names = get_source_sheet_names(workbook, template_sheet_name)

    created_sheets = []
    copy_sheets = []
    rewritten_product_names_count = 0
    updated_skus_count = 0
    existing_skus = get_existing_skus(workbook)

    for source_sheet_name in source_sheet_names:
        preferred_name = build_copy_sheet_name(source_sheet_name)
        target_sheet_name = get_unique_sheet_name(workbook, preferred_name)

        copied_sheet = workbook.copy_worksheet(template_sheet)
        copied_sheet.title = target_sheet_name
        created_sheets.append(target_sheet_name)
        copy_sheets.append(copied_sheet)
        updated_skus_count += update_missing_skus(copied_sheet, existing_skus)

    workbook.save(output_file_path)

    return workbook, {
        "source_file_path": str(source_file_path),
        "output_file_path": str(output_file_path),
        "template_sheet": template_sheet_name,
        "source_sheets": source_sheet_names,
        "created_sheets": created_sheets,
        "deleted_sheets": deleted_sheets,
        "rewritten_product_names_count": rewritten_product_names_count,
        "updated_skus_count": updated_skus_count,
    }


def build_sheet_row_maps(copy_sheets):
    sheet_row_maps = []
    unique_product_names = []
    seen_product_names = set()

    for sheet in copy_sheets:
        row_map = {}
        for row_index, product_name_column, product_name in get_product_name_rows(sheet):
            row_map.setdefault(product_name, []).append((row_index, product_name_column))
            if product_name not in seen_product_names:
                seen_product_names.add(product_name)
                unique_product_names.append(product_name)
        sheet_row_maps.append(row_map)

    return unique_product_names, sheet_row_maps


def load_unique_product_names_from_template(source_file_path):
    workbook = load_workbook(source_file_path, read_only=True, data_only=False)
    template_sheet = workbook.worksheets[-1]
    product_name_column = find_header_column(template_sheet, PRODUCT_NAME_HEADER)
    if not product_name_column:
        return []

    unique_product_names = []
    seen_product_names = set()
    product_name_index = product_name_column - 1

    for row in template_sheet.iter_rows(min_row=2, values_only=True):
        product_name = row[product_name_index]
        if not product_name or not str(product_name).strip():
            continue
        normalized_name = str(product_name).strip()
        if normalized_name not in seen_product_names:
            seen_product_names.add(normalized_name)
            unique_product_names.append(normalized_name)

    return unique_product_names


def preview_rewritten_product_names(source_file_path, model, api_key_file_path, limit_products, version_count, batch_size):
    api_key = get_openai_api_key(api_key_file_path)
    if not api_key:
        raise RuntimeError(
            "Missing OpenAI API key. Set OPENAI_API_KEY or put the key in "
            f"{DEFAULT_API_KEY_FILE_PATH}."
        )

    unique_product_names = load_unique_product_names_from_template(source_file_path)
    if limit_products is not None:
        unique_product_names = unique_product_names[:limit_products]

    parsed_structures = {}
    print(
        f"Preview parse + rewrite {len(unique_product_names)} ten san pham, moi ten {version_count} version.",
        flush=True,
    )

    for start_index in range(0, len(unique_product_names), batch_size):
        batch = unique_product_names[start_index : start_index + batch_size]
        print(
            f"Preview parse batch: {start_index + 1}-{start_index + len(batch)}/{len(unique_product_names)}",
            flush=True,
        )
        parsed_batch = request_product_name_structures_with_split(
            batch,
            model,
            api_key,
        )
        for product_name, parsed_structure in zip(batch, parsed_batch):
            parsed_structures[product_name] = parsed_structure

        parsed_products = [
            {
                "index": index,
                **parsed_structures[product_name],
            }
            for index, product_name in enumerate(batch)
        ]
        print(
            f"Preview rewrite batch: {start_index + 1}-{start_index + len(batch)}/{len(unique_product_names)}",
            flush=True,
        )
        version_batches = request_rewritten_name_versions_with_split(
            parsed_products,
            version_count,
            model,
            api_key,
        )

        for product_name, parsed_structure, versions in zip(batch, parsed_batch, version_batches):
            print("\nORIGINAL:")
            print(product_name)
            print("PARSED:")
            print(json.dumps(parsed_structure, ensure_ascii=False))
            for version_index, rewritten_name in enumerate(versions, start=1):
                print(f"VERSION {version_index}:")
                print(rewritten_name)


def rewrite_and_fill_product_names(
    workbook,
    output_file_path,
    copy_sheets,
    model,
    api_key,
    batch_size,
    processed_product_names,
    parsed_structures,
    limit_products,
):
    unique_product_names, sheet_row_maps = build_sheet_row_maps(copy_sheets)
    if limit_products is not None:
        unique_product_names = unique_product_names[:limit_products]
    remaining_product_names = [
        product_name for product_name in unique_product_names if product_name not in processed_product_names
    ]
    progress_file_path = get_progress_file_path(output_file_path)
    total_unique_product_names = len(unique_product_names)

    print(
        f"Bat dau rewrite {len(remaining_product_names)}/{total_unique_product_names} ten san pham unique con lai thanh {len(copy_sheets)} version moi ten...",
        flush=True,
    )

    rewritten_count = 0
    since_last_save = 0
    version_count = len(copy_sheets)

    for start_index in range(0, len(remaining_product_names), batch_size):
        batch = remaining_product_names[start_index : start_index + batch_size]
        missing_parsed_names = [product_name for product_name in batch if product_name not in parsed_structures]
        if missing_parsed_names:
            parsed_batch = request_product_name_structures_with_split(
                missing_parsed_names,
                model,
                api_key,
            )
            for product_name, parsed_structure in zip(missing_parsed_names, parsed_batch):
                parsed_structures[product_name] = parsed_structure

        parsed_products = [
            {
                "index": index,
                **parsed_structures[product_name],
            }
            for index, product_name in enumerate(batch)
        ]
        version_batches = request_rewritten_name_versions_with_split(
            parsed_products,
            version_count,
            model,
            api_key,
        )

        for product_name, versions in zip(batch, version_batches):
            for sheet_index, sheet in enumerate(copy_sheets):
                for row_index, product_name_column in sheet_row_maps[sheet_index].get(product_name, []):
                    sheet.cell(row=row_index, column=product_name_column).value = versions[sheet_index]
            processed_product_names.add(product_name)
            rewritten_count += 1
            since_last_save += 1

        print(
            f"Done rewrite: {len(processed_product_names)}/{total_unique_product_names}",
            flush=True,
        )

        if since_last_save >= SAVE_EVERY_UNIQUE_NAMES:
            workbook.save(output_file_path)
            save_progress(progress_file_path, processed_product_names, parsed_structures)
            print(
                f"Da save checkpoint: {len(processed_product_names)}/{total_unique_product_names}",
                flush=True,
            )
            since_last_save = 0

    if since_last_save > 0:
        workbook.save(output_file_path)
        save_progress(progress_file_path, processed_product_names, parsed_structures)
        print(
            f"Da save checkpoint: {len(processed_product_names)}/{total_unique_product_names}",
            flush=True,
        )

    return rewritten_count


def create_copy_workbook(source_file_path, output_file_path, model, batch_size, skip_ai, api_key_file_path, limit_products):
    progress_file_path = get_progress_file_path(output_file_path)
    api_key = get_openai_api_key(api_key_file_path)

    if not skip_ai and not api_key:
        raise RuntimeError(
            "Missing OpenAI API key. Set OPENAI_API_KEY or put the key in "
            f"{DEFAULT_API_KEY_FILE_PATH}."
        )

    if output_file_path.exists() and progress_file_path.exists():
        workbook = load_workbook(output_file_path)
        template_sheet_name = get_template_sheet_name(workbook)
        source_sheet_names = get_source_sheet_names(workbook, template_sheet_name)
        created_sheets = get_copy_sheet_names_from_source_sheet_names(source_sheet_names)
        copy_sheets = [workbook[sheet_name] for sheet_name in created_sheets if sheet_name in workbook.sheetnames]
        progress = load_progress(progress_file_path)
        processed_product_names = set(progress.get("processed_product_names", []))
        parsed_structures = {
            product_name: normalize_product_name_structure(product_name, structure)
            for product_name, structure in progress.get("parsed_structures", {}).items()
        }
        updated_skus_count = 0
        deleted_sheets = []
        print(
            f"Tiep tuc tu checkpoint: {len(processed_product_names)} ten da xu ly.",
            flush=True,
        )
    else:
        workbook, result = create_initial_copy_workbook(source_file_path, output_file_path)
        template_sheet_name = result["template_sheet"]
        source_sheet_names = result["source_sheets"]
        created_sheets = result["created_sheets"]
        deleted_sheets = result["deleted_sheets"]
        updated_skus_count = result["updated_skus_count"]
        copy_sheets = [workbook[sheet_name] for sheet_name in created_sheets]
        processed_product_names = set()
        parsed_structures = {}
        save_progress(progress_file_path, processed_product_names, parsed_structures)

    rewritten_product_names_count = 0
    if not skip_ai and copy_sheets:
        rewritten_product_names_count = rewrite_and_fill_product_names(
            workbook,
            output_file_path,
            copy_sheets,
            model,
            api_key,
            batch_size,
            processed_product_names,
            parsed_structures,
            limit_products,
        )

    workbook.save(output_file_path)
    save_progress(progress_file_path, processed_product_names, parsed_structures)

    return {
        "source_file_path": str(source_file_path),
        "output_file_path": str(output_file_path),
        "template_sheet": template_sheet_name,
        "source_sheets": source_sheet_names,
        "created_sheets": created_sheets,
        "deleted_sheets": deleted_sheets,
        "rewritten_product_names_count": rewritten_product_names_count,
        "updated_skus_count": updated_skus_count,
        "processed_product_names_count": len(processed_product_names),
    }


def parse_args():
    parser = argparse.ArgumentParser(
        description="Copy the last worksheet as a template for every existing data worksheet."
    )
    parser.add_argument(
        "source_file_path",
        nargs="?",
        default=DEFAULT_FILE_PATH,
        type=Path,
        help=f"Source Excel file path. Default: {DEFAULT_FILE_PATH}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE_PATH,
        type=Path,
        help=f"Output Excel file path. Default: {DEFAULT_OUTPUT_FILE_PATH}",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"OpenAI model used to rewrite product names. Default: {DEFAULT_MODEL}",
    )
    parser.add_argument(
        "--batch-size",
        default=40,
        type=int,
        help="Number of product names to send in each OpenAI request.",
    )
    parser.add_argument(
        "--skip-ai",
        action="store_true",
        help="Create copy sheets and SKUs without rewriting product names.",
    )
    parser.add_argument(
        "--api-key-file",
        default=DEFAULT_API_KEY_FILE_PATH,
        type=Path,
        help=f"File containing OpenAI API key. Default: {DEFAULT_API_KEY_FILE_PATH}",
    )
    parser.add_argument(
        "--limit-products",
        type=int,
        default=None,
        help="Only process the first N unique product names. Useful for testing.",
    )
    parser.add_argument(
        "--preview-only",
        action="store_true",
        help="Parse and rewrite a small sample, then print to terminal without writing any file.",
    )
    parser.add_argument(
        "--preview-version-count",
        type=int,
        default=3,
        help="Number of rewritten versions per product in preview mode.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    if args.preview_only:
        preview_rewritten_product_names(
            args.source_file_path,
            args.model,
            args.api_key_file,
            args.limit_products,
            args.preview_version_count,
            args.batch_size,
        )
        return

    result = create_copy_workbook(
        args.source_file_path,
        args.output,
        model=args.model,
        batch_size=args.batch_size,
        skip_ai=args.skip_ai,
        api_key_file_path=args.api_key_file,
        limit_products=args.limit_products,
    )

    print(f"File goc: {result['source_file_path']}")
    print(f"File moi: {result['output_file_path']}")
    print(f"Sheet template cuoi cung: {result['template_sheet']}")
    print(f"So sheet copy da xoa trong file moi: {len(result['deleted_sheets'])}")
    print(f"So sheet nguon: {len(result['source_sheets'])}")
    print(f"So sheet da tao: {len(result['created_sheets'])}")
    print(f"So ten san pham da doi: {result['rewritten_product_names_count']}")
    print(f"So SKU moi da them: {result['updated_skus_count']}")

    if result["created_sheets"]:
        print("Sheet moi:")
        for sheet_name in result["created_sheets"]:
            print(f"  - {sheet_name}")

    if result["deleted_sheets"]:
        print("Sheet copy da xoa trong file moi:")
        for sheet_name in result["deleted_sheets"]:
            print(f"  - {sheet_name}")


if __name__ == "__main__":
    main()
