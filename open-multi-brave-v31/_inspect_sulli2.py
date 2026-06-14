import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
wb = load_workbook(r"D:\Projects\shopee\data\data.xlsx", read_only=True, data_only=True)
ws = wb["sulli"]
has_f_rows = []
for r in range(1056, 2056):
    f = str(ws.cell(r, 6).value or "").strip()
    d = str(ws.cell(r, 4).value or "").strip()
    g = str(ws.cell(r, 7).value or "").strip()
    if f:
        has_f_rows.append((r, d[:20], g[:20], f[:40]))
print("rows with F from 1056:", len(has_f_rows))
for item in has_f_rows[:15]:
    print(item)
# rows with only link
link_only = 0
for r in range(1056, 2056):
    a = str(ws.cell(r, 1).value or "").strip()
    f = str(ws.cell(r, 6).value or "").strip()
    if a and not f:
        link_only += 1
print("link-only rows 1056-2055:", link_only)
# last rows with data before 1056
for r in range(1040, 1056):
    f = str(ws.cell(r, 6).value or "").strip()
    g = str(ws.cell(r, 7).value or "").strip()
    if f:
        print(f"row {r} F=yes G={'yes' if g else 'no'}")
wb.close()
