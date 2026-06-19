"""Workbook, XLSX lookup, and row-marking helpers for product updates."""
import json
import os
import re
import shutil
import hashlib
import threading
import time
import zipfile
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from product_runtime import log

try:
    import fcntl
except ImportError:
    fcntl = None

try:
    import msvcrt
except ImportError:
    msvcrt = None

_WORKBOOK_LOCK = threading.Lock()
_WORKBOOK_INDEX_LOCK = threading.Lock()
_WORKBOOK_INDEX_CACHE = {}  # (workbook_path, sheet_name) -> {"index": {id: result}, "built_from_row": int}

def _workbook_lock_path(workbook_path):
    return Path(f"{Path(workbook_path)}.lock")


def _acquire_windows_file_lock(lock_file, timeout_sec=120):
    deadline = time.time() + timeout_sec
    while True:
        try:
            lock_file.seek(0)
            msvcrt.locking(lock_file.fileno(), msvcrt.LK_LOCK, 1)
            return
        except OSError as exc:
            if time.time() >= deadline:
                raise TimeoutError(f"Khong acquire duoc workbook lock sau {timeout_sec}s") from exc
            time.sleep(0.1)


def _release_windows_file_lock(lock_file):
    lock_file.seek(0)
    msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)


@contextmanager
def workbook_file_lock(workbook_path):
    workbook_path = Path(workbook_path)
    lock_path = _workbook_lock_path(workbook_path)
    lock_path.parent.mkdir(parents=True, exist_ok=True)

    if fcntl is not None:
        with lock_path.open("w") as lock_file:
            fcntl.flock(lock_file, fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(lock_file, fcntl.LOCK_UN)
        return

    if msvcrt is not None:
        lock_file = None
        try:
            deadline = time.time() + 120
            while lock_file is None:
                try:
                    lock_file = lock_path.open("a+b")
                except PermissionError as exc:
                    if time.time() >= deadline:
                        raise TimeoutError(
                            f"Khong mo duoc workbook lock file: {lock_path}"
                        ) from exc
                    time.sleep(0.1)
            _acquire_windows_file_lock(lock_file)
            try:
                yield
            finally:
                _release_windows_file_lock(lock_file)
        finally:
            if lock_file is not None:
                lock_file.close()
        return

    _WORKBOOK_LOCK.acquire()
    try:
        yield
    finally:
        _WORKBOOK_LOCK.release()


def _is_valid_xlsx(path):
    try:
        with zipfile.ZipFile(path, "r") as archive:
            return archive.testzip() is None
    except (OSError, zipfile.BadZipFile):
        return False


def load_workbook_checked(workbook_path, *args, **kwargs):
    workbook_path = Path(workbook_path)
    deadline = time.time() + 120
    last_permission_error = None
    while time.time() < deadline:
        try:
            return load_workbook(workbook_path, *args, **kwargs)
        except PermissionError as exc:
            last_permission_error = exc
            time.sleep(0.1)
        except zipfile.BadZipFile as exc:
            return _raise_bad_workbook_error(workbook_path, exc)

    if last_permission_error is not None:
        raise PermissionError(
            f"Khong mo duoc workbook (file dang bi khoa): {workbook_path}"
        ) from last_permission_error

    try:
        return load_workbook(workbook_path, *args, **kwargs)
    except PermissionError as exc:
        raise PermissionError(
            f"Khong mo duoc workbook (file dang bi khoa): {workbook_path}"
        ) from exc
    except zipfile.BadZipFile as exc:
        return _raise_bad_workbook_error(workbook_path, exc)


def _raise_bad_workbook_error(workbook_path, exc):
    backup_path = Path(f"{workbook_path}.bak")
    if backup_path.exists() and _is_valid_xlsx(backup_path):
        raise RuntimeError(
            f"Workbook bi hong: {workbook_path}. Co backup hop le: {backup_path}. "
            "Hay doi file backup thanh workbook chinh roi chay lai."
        ) from exc
    raise RuntimeError(
        f"Workbook bi hong/ghi do dang: {workbook_path}. "
        "Can khoi phuc file .xlsx tu backup hoac file goc roi chay lai."
    ) from exc


def save_workbook_atomic(workbook, workbook_path):
    workbook_path = Path(workbook_path)
    temp_path = workbook_path.with_name(
        f".{workbook_path.name}.{os.getpid()}.{threading.get_ident()}.tmp"
    )
    backup_path = Path(f"{workbook_path}.bak")

    try:
        workbook.save(temp_path)
        if not _is_valid_xlsx(temp_path):
            raise RuntimeError(f"File tam sau khi save khong phai XLSX hop le: {temp_path}")

        if workbook_path.exists() and _is_valid_xlsx(workbook_path):
            shutil.copy2(workbook_path, backup_path)

        try:
            os.replace(temp_path, workbook_path)
        except PermissionError:
            log(f"Atomic replace bi chan, save truc tiep vao workbook: {workbook_path}")
            workbook.save(workbook_path)
            if not _is_valid_xlsx(workbook_path):
                raise RuntimeError(f"Workbook sau khi save truc tiep khong hop le: {workbook_path}")
    finally:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except OSError:
            pass


def get_cell_fill_color(cell):
    color = cell.fill.fgColor
    return str(color.rgb or "").upper() if color and color.type == "rgb" else ""


def is_scraped_ok_row(sheet, row_index):
    return any(
        get_cell_fill_color(sheet.cell(row=row_index, column=column_index)) == SCRAPED_OK_FILL_COLOR
        for column_index in range(1, sheet.max_column + 1)
    )


def mark_row_fill(sheet, row_index, fill_color, preserve_scraped=True):
    if preserve_scraped and is_scraped_ok_row(sheet, row_index):
        return False

    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    for column_index in range(1, sheet.max_column + 1):
        sheet.cell(row=row_index, column=column_index).fill = fill
    return True


def mark_workbook_row(workbook_path, sheet_name, row_index, fill_color):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        log(f"⚠️ Không tìm thấy workbook để tô màu: {workbook_path}")
        return False

    with workbook_file_lock(workbook_path):
        workbook = load_workbook_checked(workbook_path)
        sheet = workbook[resolve_sheet_name(workbook, sheet_name)]
        updated = mark_row_fill(sheet, int(row_index), fill_color)
        if updated:
            save_workbook_atomic(workbook, workbook_path)
    return updated

def parse_price(price_str):
    """Chuyển giá dạng '304k' hoặc '304000' → số nguyên"""
    s = str(price_str).strip().lower().replace(',', '').replace(' ', '')
    if s.endswith('k'):
        try:
            return int(float(s[:-1]) * 1000)
        except:
            pass
    digits = ''.join(filter(str.isdigit, s))
    return int(digits) if digits else 0

SHOPEE_BLOCKED_URL_MARKERS = ("/verify/captcha", "/verify/traffic")
SHOPEE_BLOCKED_MARKER = "__SHOPEE_BLOCKED__"


def is_shopee_blocked_url(url):
    u = (url or "").strip().lower()
    return any(marker in u for marker in SHOPEE_BLOCKED_URL_MARKERS)


def extract_shopee_id(url):
    """Cắt Shopee item ID từ link Shopee (sau nút Chép) — KHÔNG dùng URL tab edit BigSeller."""
    if is_shopee_blocked_url(url):
        return None
    try:
        # Dạng 1: ...-i.ShopID.ItemID
        match = re.search(r"i\.\d+\.(\d+)", url)
        if match: return match.group(1)
        
        # Dạng 2: .../product/ShopID/ItemID
        match2 = re.search(r"product\/\d+\/(\d+)", url)
        if match2: return match2.group(1)

        # Dạng 3: Link rút gọn hoặc query param
        match3 = re.search(r"i\.\d+\.(\d+)\?", url)
        if match3: return match3.group(1)
    except: pass
    return None

PRODUCT_LINK_HEADER = "link sp"
# Ưu tiên giá bán (cột C) rồi mới fallback giá gốc (cột B).
# Giữ tương thích header cũ "giá bán (giá gốc +150k)".
PRICE_HEADERS = ("gi\u00e1 b\u00e1n", "gi\u00e1 b\u00e1n (gi\u00e1 g\u1ed1c +150k)", "gi\u00e1 g\u1ed1c")
SKU_HEADER = "SKU"
SOURCE_PRODUCT_ID_HEADER = "ID s\u1ea3n ph\u1ea9m g\u1ed1c"
PRODUCT_NAME_HEADER = "T\u00ean sp"
REWRITTEN_PRODUCT_NAME_HEADER = "T\u00ean sp \u0111\u00e3 s\u1eeda"
# Fixed column layout (row 1 = header). Update product reads by position, not header name.
# A=link | B=original price | C=sale price | D=SKU | E=source product ID | F=product name | G=rewritten name
DATA_LINK_COLUMN = 1
DATA_ORIGINAL_PRICE_COLUMN = 2
DATA_SALE_PRICE_COLUMN = 3
DATA_SKU_COLUMN = 4
DATA_SOURCE_ID_COLUMN = 5
DATA_PRODUCT_NAME_COLUMN = 6
DATA_REWRITTEN_NAME_COLUMN = 7
MISSING_REWRITTEN_NAME_FILL_COLOR = "FF00B0F0"
REWRITTEN_NAME_OK_FILL_COLOR = "FF00B050"
SCRAPED_OK_FILL_COLOR = "FF800080"
ATTRIBUTE_START_PATTERNS = [
    ("cao",),
    ("êm", "chân"),
    ("da", "mềm"),
    ("đính", "nơ"),
    ("dây", "cài"),
    ("quai", "cài"),
    ("may", "viền"),
    ("hot", "trend"),
    ("dễ", "phối"),
    ("tôn", "dáng"),
    ("phong", "cách"),
]

DESCRIPTION_CLEANUP_PATTERNS = [
    r"\bdễ phối(?: đồ| trang phục| outfit)?\b",
    r"\bde phoi(?: do| trang phuc| outfit)?\b",
    r"\bdễ dàng phối(?: đồ| trang phục| outfit)?\b",
    r"\bde dang phoi(?: do| trang phuc| outfit)?\b",
    r"\bhằng ngày\b",
    r"\bhang ngay\b",
    r"\bphong cách\b",
    r"\bphong cach\b",
    r"\bthanh lịch\b",
    r"\bthanh lich\b",
    r"\bnhẹ nhàng\b",
    r"\bnhe nhang\b",
    r"\bêm ái\b",
    r"\bem ai\b",
    r"\bkiểu dáng\b",
    r"\bkieu dang\b",
    r"\bdáng vẻ\b",
    r"\bdang ve\b",
    r"\blựa chọn tuyệt vời\b",
    r"\blua chon tuyet voi\b",
    r"\bhoàn hảo\b",
    r"\bhoan hao\b",
    r"\bmọi dịp\b",
    r"\bmoi dip\b",
    r"\bmang lại sự\b",
    r"\bmang lai su\b",
    r"\btiện dụng\b",
    r"\btien dung\b",
    r"\bdễ kết hợp\b",
    r"\bde ket hop\b",
]
MAX_KEYWORD_2_WORDS = 3


def normalize_text(value):
    return str(value or "").strip().lower()


def find_header_column(sheet, header_name):
    normalized_header_name = normalize_text(header_name)
    for column_index, cell in enumerate(sheet[1], start=1):
        if normalize_text(cell.value) == normalized_header_name:
            return column_index
    return None


def find_first_header_column(sheet, header_names):
    for header_name in header_names:
        column = find_header_column(sheet, header_name)
        if column:
            return column
    return None


def get_or_create_header_column(sheet, header_name):
    column = find_header_column(sheet, header_name)
    if column:
        return column
    column = sheet.max_column + 1
    sheet.cell(row=1, column=column).value = header_name
    return column


def get_or_create_rewritten_name_column(sheet, product_name_column=None):
    """Always use fixed column G for rewritten product names."""
    _ = product_name_column
    column = DATA_REWRITTEN_NAME_COLUMN
    if not normalize_text(sheet.cell(row=1, column=column).value):
        sheet.cell(row=1, column=column).value = REWRITTEN_PRODUCT_NAME_HEADER
    return column


def resolve_sheet_name(workbook, sheet_name):
    normalized_sheet_name = normalize_text(sheet_name)
    for candidate in workbook.sheetnames:
        if normalize_text(candidate) == normalized_sheet_name:
            return candidate
    raise ValueError(f"Không tìm thấy sheet: {sheet_name}")

def _resolve_data_row_range(sheet, start_row=2, end_row=None):
    first_data_row = max(2, int(start_row or 2))
    last_sheet_row = sheet.max_row or first_data_row
    last_data_row = min(last_sheet_row, int(end_row)) if end_row else last_sheet_row
    return first_data_row, max(first_data_row, last_data_row)


def _batch_cache_dir(workbook_path):
    root = Path(os.environ.get("BIGSELLER_PROFILE_DIR") or Path(workbook_path).parent)
    cache_dir = root.parent / ".update-product-workbook-cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def _batch_cache_key(workbook_path, sheet_name, start_row, end_row):
    batch_id = os.environ.get("BIGSELLER_UPDATE_BATCH_ID", "").strip()
    scope = "|".join([
        batch_id,
        str(Path(workbook_path).resolve()),
        normalize_text(sheet_name),
        str(int(start_row or 2)),
        str(int(end_row or 0)),
    ])
    return hashlib.sha1(scope.encode("utf-8", errors="ignore")).hexdigest()[:24]


@contextmanager
def _json_file_lock(lock_path, timeout_sec=120):
    lock_path = Path(lock_path)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    deadline = time.time() + timeout_sec
    handle = None
    while handle is None:
        try:
            handle = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError as exc:
            try:
                if time.time() - lock_path.stat().st_mtime > timeout_sec:
                    lock_path.unlink()
            except Exception:
                pass
            if time.time() >= deadline:
                raise TimeoutError(f"Khong lay duoc workbook cache lock: {lock_path}") from exc
            time.sleep(0.1)
    try:
        yield
    finally:
        try:
            os.close(handle)
        except Exception:
            pass
        try:
            lock_path.unlink()
        except Exception:
            pass


def _batch_cache_path(workbook_path, sheet_name, start_row, end_row):
    key = _batch_cache_key(workbook_path, sheet_name, start_row, end_row)
    return _batch_cache_dir(workbook_path) / f"{key}.json"


def _load_batch_workbook_cache(workbook_path, sheet_name, start_row, end_row):
    workbook_path = Path(workbook_path)
    cache_path = _batch_cache_path(workbook_path, sheet_name, start_row, end_row)
    lock_path = cache_path.with_suffix(".lock")

    with _json_file_lock(lock_path):
        if cache_path.exists():
            try:
                with cache_path.open("r", encoding="utf-8") as handle:
                    payload = json.load(handle)
                if isinstance(payload, dict) and isinstance(payload.get("index"), dict):
                    return payload
            except Exception:
                pass

        built = _build_workbook_index(workbook_path, sheet_name, start_row, end_row)
        payload = {
            "built_at": time.time(),
            "workbook_path": str(workbook_path),
            "sheet_name": sheet_name,
            "start_row": int(start_row or 2),
            "end_row": int(end_row or 0),
            "index": built.get("index") or {},
            "missing_product_name_index": built.get("missing_product_name_index") or {},
        }
        temp_path = cache_path.with_suffix(f".{os.getpid()}.tmp")
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False)
        os.replace(temp_path, cache_path)
        return payload


def search_in_workbook(target_id, workbook_path, sheet_name, start_row=2, end_row=None):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        log(f"⚠️ Không tìm thấy workbook: {workbook_path}")
        return None

    # Cache/index để lookup nhanh theo ID trong quá trình update.
    # Mặc định bật; có thể tắt bằng env SHOPEE_DISABLE_WORKBOOK_CACHE=1
    target_id = str(target_id or "").strip()
    if os.environ.get("SHOPEE_DISABLE_WORKBOOK_CACHE", "0") != "1":
        try:
            cached = _load_batch_workbook_cache(workbook_path, sheet_name, start_row, end_row)
            index = cached.get("index") or {}
            missing_product_name_index = cached.get("missing_product_name_index") or {}
            return index.get(target_id) or missing_product_name_index.get(target_id)
        except Exception as cache_error:
            log(f"Khong tai duoc workbook cache, fallback doc XLSX: {cache_error}")

    with workbook_file_lock(workbook_path):
        workbook = load_workbook_checked(workbook_path)
        sheet = workbook[resolve_sheet_name(workbook, sheet_name)]
        link_column = DATA_LINK_COLUMN
        price_columns = [DATA_SALE_PRICE_COLUMN]
        price_columns = [column for column in price_columns if column]
        sku_column = DATA_SKU_COLUMN
        source_id_column = DATA_SOURCE_ID_COLUMN
        product_name_column = DATA_PRODUCT_NAME_COLUMN
        rewritten_name_column = DATA_REWRITTEN_NAME_COLUMN


        first_data_row, last_data_row = _resolve_data_row_range(sheet, start_row, end_row)
        log(f"🔎 Quét XLSX sheet '{sheet.title}' từ row {first_data_row}-{last_data_row} tìm ID: {target_id}...")
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, values_only=True),
            start=first_data_row,
        ):
            if str(target_id).strip() not in str(row):
                continue

            link = row[link_column - 1] if len(row) >= link_column else ""
            source_id = row[source_id_column - 1] if source_id_column and len(row) >= source_id_column else ""
            row_id = str(source_id or "").strip() or extract_shopee_id(link)
            if str(row_id or "").strip() != target_id:
                continue

            price = ""
            for price_column in price_columns:
                value = row[price_column - 1] if len(row) >= price_column else ""
                if value is not None and str(value).strip():
                    price = str(value).strip()
                    break

            sku = str(row[sku_column - 1] or "").strip()
            original_name = str(row[product_name_column - 1] or "").strip()
            rewritten_name = str(row[rewritten_name_column - 1] or "").strip() if len(row) >= rewritten_name_column else ""
            update_name = rewritten_name or original_name
            if not update_name:
                log(f"Row {row_index} has ID {target_id} but both column F and G are empty -> skip.")
                return {
                    "status": "missing_product_name",
                    "line_index": row_index,
                    "source_product_id": row_id,
                    "sheet": sheet.title,
                }
            return {
                "line_index": row_index,
                "link": str(link or "").strip(),
                "price": price,
                "sku": sku,
                "source_product_id": row_id,
                "original_product_name": original_name,
                "product_name": update_name,
                "has_rewritten_product_name": bool(rewritten_name),
                "sheet": sheet.title,
            }

    return None


def _build_workbook_index(workbook_path: Path, sheet_name: str, start_row: int, end_row=None):
    workbook_path = Path(workbook_path)
    # Use a full scan once, then O(1) lookup for each product.
    with workbook_file_lock(workbook_path):
        workbook = load_workbook_checked(workbook_path, read_only=True, data_only=True)
        sheet = workbook[resolve_sheet_name(workbook, sheet_name)]

        link_column = DATA_LINK_COLUMN
        price_columns = [DATA_SALE_PRICE_COLUMN]
        price_columns = [column for column in price_columns if column]
        sku_column = DATA_SKU_COLUMN
        source_id_column = DATA_SOURCE_ID_COLUMN
        product_name_column = DATA_PRODUCT_NAME_COLUMN
        rewritten_name_column = DATA_REWRITTEN_NAME_COLUMN


        first_data_row, last_data_row = _resolve_data_row_range(sheet, start_row, end_row)
        log(f"🧠 Build workbook index sheet '{sheet.title}' từ row {first_data_row}-{last_data_row}...")

        index = {}
        missing_product_name_index = {}
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, values_only=True),
            start=first_data_row,
        ):
            link = row[link_column - 1] if len(row) >= link_column else ""
            source_id = row[source_id_column - 1] if source_id_column and len(row) >= source_id_column else ""
            row_id = str(source_id or "").strip() or extract_shopee_id(link)
            if not row_id:
                continue

            price = ""
            for price_column in price_columns:
                value = row[price_column - 1] if len(row) >= price_column else ""
                if value is not None and str(value).strip():
                    price = str(value).strip()
                    break

            sku = str(row[sku_column - 1] or "").strip()
            original_name = str(row[product_name_column - 1] or "").strip()
            rewritten_name = str(row[rewritten_name_column - 1] or "").strip() if len(row) >= rewritten_name_column else ""
            update_name = rewritten_name or original_name
            if not update_name:
                missing_product_name_index[str(row_id).strip()] = {
                    "status": "missing_product_name",
                    "line_index": row_index,
                    "source_product_id": str(row_id).strip(),
                    "sheet": sheet.title,
                }
                continue
            index[str(row_id).strip()] = {
                "line_index": row_index,
                "link": str(link or "").strip(),
                "price": price,
                "sku": sku,
                "source_product_id": str(row_id).strip(),
                "original_product_name": original_name,
                "product_name": update_name,
                "has_rewritten_product_name": bool(rewritten_name),
                "sheet": sheet.title,
            }

        return {
            "index": index,
            "missing_product_name_index": missing_product_name_index,
            "built_from_row": first_data_row,
        }


def _workbook_mtime(workbook_path):
    workbook_path = Path(workbook_path)
    try:
        return workbook_path.stat().st_mtime_ns
    except OSError:
        return None
